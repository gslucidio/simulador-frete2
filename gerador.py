"""
Gerador de Modelo de Frete
Informe o número de dias, trânsito e os 3 prazos do importador.
Todos os outros parâmetros são editados direto no Excel (Dashboard).
"""

import io
import datetime
import streamlit as st
from openpyxl import load_workbook
from openpyxl.cell import MergedCell
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Gerador — Modelo de Frete", page_icon="🚛")
st.title("🚛 Gerador — Modelo de Frete")
st.divider()

col1, col2 = st.columns(2)

with col1:
    n_dias   = st.number_input("Número de dias", min_value=50, max_value=1800, value=360, step=10)
    transito = st.number_input("Dias de trânsito (corridos)", min_value=1, max_value=60, value=5, step=1)

with col2:
    d1 = st.number_input("Prazo importador cart. 1 (dias)", min_value=1, max_value=365, value=10, step=1)
    d2 = st.number_input("Prazo importador cart. 2 (dias)", min_value=1, max_value=365, value=12, step=1)
    d3 = st.number_input("Prazo importador cart. 3 (dias)", min_value=1, max_value=365, value=14, step=1)
    if not (d1 <= d2 <= d3):
        st.warning("Os prazos devem estar em ordem crescente (d1 ≤ d2 ≤ d3).")

st.divider()


# ════════════════════════════════════════════════════════════════════════════
# GERADOR
# ════════════════════════════════════════════════════════════════════════════

def gerar_excel(template_bytes, n_dias, transito, d1, d2, d3):
    N     = n_dias
    LR    = N + 4
    START = datetime.datetime(2026, 6, 1)  # segunda-feira

    def is_wd(n):
        return (n - 1) % 7 <= 4

    def next_wd(n):
        while not is_wd(n):
            n += 1
        return n

    # ── ORIG: último dia útil em que op ainda fecha dentro de N ────────────
    ORIG = 0
    for n0 in range(1, N + 1):
        if not is_wd(n0):
            continue
        dt  = next_wd(n0 + transito)
        if dt > N:
            break
        dv3 = next_wd(dt + d3)
        if dv3 > N:
            break
        ORIG = n0

    wb   = load_workbook(io.BytesIO(template_bytes))
    ws_f = wb["Fundo"]
    ws_c = wb["Custos"]
    ws_d = wb["Dashboard"]

    def sc(ws, row, col, val):
        cell = ws.cell(row, col)
        if not isinstance(cell, MergedCell):
            cell.value = val

    def clear_rows(ws, from_row):
        for r in range(from_row, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                sc(ws, r, col, None)

    # ── Dashboard ──────────────────────────────────────────────────────────
    ws_d["I7"].value  = transito
    ws_d["I8"].value  = ORIG
    ws_d["L14"].value = d1
    ws_d["L15"].value = d2
    ws_d["L16"].value = d3

    # Refs dinâmicas à última linha
    ws_d["E5"].value  = f"=SUM(Fundo!AG5:AG{LR})"
    ws_d["E6"].value  = f"=SUM(Fundo!AM5:AM{LR})"
    ws_d["E7"].value  = f"=SUM(Fundo!AQ5:AQ{LR})"
    ws_d["C23"].value = f"=Fundo!X{LR}/Dashboard!D14"

    # ── Fundo: cabeçalhos ──────────────────────────────────────────────────
    ws_f["E2"].value  = f"Carteira {d1}d"
    ws_f["J2"].value  = f"Carteira {d2}d"
    ws_f["O2"].value  = f"Carteira {d3}d"

    # IRRs
    ws_f["AA2"].value = f"=IRR(AA5:AA{LR},0.001)"
    ws_f["AG2"].value = f"=IRR(AG5:AG{LR},0.001)"
    ws_f["AM2"].value = f"=IRR(AM5:AM{LR},0.001)"
    ws_f["AQ2"].value = f"=IRR(AQ5:AQ{LR},0.001)"

    # Linha 4 (init)
    ws_f["X4"].value  = "=AC5+AI5+AO5"
    ws_f["Z4"].value  = "=X4"
    ws_f["AC4"].value = "=Dashboard!D10"
    ws_f["AI4"].value = "=Dashboard!D11"
    ws_f["AO4"].value = "=Dashboard!D12"

    # ── Limpar dados antigos ───────────────────────────────────────────────
    clear_rows(ws_f, 5)
    clear_rows(ws_c, 5)

    yellow    = PatternFill(patternType="solid", fgColor="FFFFFFCC")
    light_red = PatternFill(patternType="solid", fgColor="FFFFCCCC")

    # Atalhos de range
    rng_at = f"$AT$5:$AT${LR}"
    rng_au = f"$AU$5:$AU${LR}"
    rng_av = f"$AV$5:$AV${LR}"
    rng_aw = f"$AW$5:$AW${LR}"
    rng_e  = f"$E$5:$E${LR}"
    rng_j  = f"$J$5:$J${LR}"
    rng_o  = f"$O$5:$O${LR}"

    # ── Fundo: linhas 5..LR ────────────────────────────────────────────────
    for n in range(1, N + 1):
        r     = n + 4
        pr    = r - 1
        first = (n == 1)
        last  = (n == N)
        wd    = is_wd(n)

        # B: Benchmark (base 252)
        sc(ws_f,r,2,  "=Dashboard!D14" if first
                      else f"=IF(WEEKDAY($D{r},2)>5,B{pr},B{pr}*((1+Dashboard!$D$20)^(1/252)))")

        # C: contador, D: data
        sc(ws_f,r,3, n)
        sc(ws_f,r,4, START if first else f"=D{pr}+1")

        # ── Originações E (5), J (10), O (15) ──────────────────────────────
        for col, dist_col in [(5, "$I$14"), (10, "$I$15"), (15, "$I$16")]:
            sc(ws_f,r,col,
               f"=IF(AND(WEEKDAY($D{r},2)<=5,C{r}<=Dashboard!$I$8),Dashboard!$I$13*Dashboard!{dist_col},0)")

        # ── 1º Desembolso F (6), K (11), P (16) ────────────────────────────
        sc(ws_f,r,6,  f"=IF(WEEKDAY($D{r},2)>5,0,E{r}*Dashboard!$J$10*Dashboard!$I$6)")
        sc(ws_f,r,11, f"=IF(WEEKDAY($D{r},2)>5,0,J{r}*Dashboard!$J$11*Dashboard!$I$6)")
        sc(ws_f,r,16, f"=IF(WEEKDAY($D{r},2)>5,0,O{r}*Dashboard!$J$12*Dashboard!$I$6)")

        # ── 2º Desembolso G (7), L (12), Q (17) ────────────────────────────
        sc(ws_f,r,7,  f"=IF(WEEKDAY($D{r},2)>5,0,SUMIF({rng_at},$D{r},{rng_e})*Dashboard!$J$10*(1-Dashboard!$I$6))")
        sc(ws_f,r,12, f"=IF(WEEKDAY($D{r},2)>5,0,SUMIF({rng_at},$D{r},{rng_j})*Dashboard!$J$11*(1-Dashboard!$I$6))")
        sc(ws_f,r,17, f"=IF(WEEKDAY($D{r},2)>5,0,SUMIF({rng_at},$D{r},{rng_o})*Dashboard!$J$12*(1-Dashboard!$I$6))")

        # ── Levantamento H (8), M (13), R (18) ─────────────────────────────
        sc(ws_f,r,8,  f"=IF(WEEKDAY($D{r},2)>5,0,SUMIF({rng_au},$D{r},{rng_e})*Dashboard!$I$4)")
        sc(ws_f,r,13, f"=IF(WEEKDAY($D{r},2)>5,0,SUMIF({rng_av},$D{r},{rng_j})*Dashboard!$I$4)")
        sc(ws_f,r,18, f"=IF(WEEKDAY($D{r},2)>5,0,SUMIF({rng_aw},$D{r},{rng_o})*Dashboard!$I$4)")

        # ── T (20): n total; U (21): Total Desemb; V (22): Levantamento ────
        sc(ws_f,r,20, f"=E{r}+J{r}+O{r}")
        sc(ws_f,r,21, f"=F{r}+G{r}+K{r}+L{r}+P{r}+Q{r}")
        sc(ws_f,r,22, f"=H{r}+M{r}+R{r}")

        # ── W (23): P&L Ativos ─────────────────────────────────────────────
        sc(ws_f,r,23, f"=V{r}-(SUMIF({rng_au},$D{r},{rng_e})*Dashboard!$J$10"
                      f"+SUMIF({rng_av},$D{r},{rng_j})*Dashboard!$J$11"
                      f"+SUMIF({rng_aw},$D{r},{rng_o})*Dashboard!$J$12)")

        # ── Y (25): Rec LM (CDI sobre caixa, base 252) ─────────────────────
        sc(ws_f,r,25, 0 if first
                      else f"=IF(WEEKDAY($D{r},2)>5,0,X{pr}*((1+Dashboard!$D$17)^(1/252)-1))")

        # ── X (24): Caixa ──────────────────────────────────────────────────
        if first:
            sc(ws_f,r,24, f"=X4-U{r}+V{r}+Y{r}-Custos!D{r}")
        elif n == 2:
            # Dia 2: integralizações iniciais já estão em X4, só desconta amortizações
            sc(ws_f,r,24, f"=X{pr}-U{r}+V{r}+Y{r}-Custos!D{r}"
                          f"-AE{pr}-AK{pr}-AP{pr}")
        else:
            sc(ws_f,r,24, f"=X{pr}-U{r}+V{r}+Y{r}-Custos!D{r}"
                          f"+AC{pr}+AI{pr}+AO{pr}-AE{pr}-AK{pr}-AP{pr}")

        # ── Z (26): PL Fundo = Caixa + Carteira ────────────────────────────
        sc(ws_f,r,26, f"=X{r}+Carteira!C{r-1}")

        # ── AA (27): Fluxo ─────────────────────────────────────────────────
        sc(ws_f,r,27, f"=-U{r}+V{r}-Custos!D{r}")

        # ── SÊNIOR (AC=29, AD=30, AE=31, AF=32, AG=33) ─────────────────────
        sc(ws_f,r,29, "=AC4" if first else 0)
        sc(ws_f,r,30, f"=IF(WEEKDAY($D{r},2)>5,0,AF{pr}*((1+Dashboard!$D$20)^(1/252)-1))")
        sc(ws_f,r,31, f"=AF{r}" if last else 0)
        sc(ws_f,r,32, f"=AC{r}" if first else f"=AF{pr}+AD{r}-AE{pr}+AC{r}")
        sc(ws_f,r,33, f"=-AC{r}+AE{r}")

        # ── MEZANINO (AI=35, AJ=36, AK=37, AL=38, AM=39) ───────────────────
        sc(ws_f,r,35, "=AI4" if first else 0)
        sc(ws_f,r,36, f"=IF(WEEKDAY($D{r},2)>5,0,AL{pr}*((1+Dashboard!$D$21)^(1/252)-1))")
        sc(ws_f,r,37, f"=AL{r}" if last else 0)
        sc(ws_f,r,38, f"=AI{r}" if first else f"=AL{pr}+AJ{r}-AK{pr}+AI{r}")
        sc(ws_f,r,39, f"=-AI{r}+AK{r}")

        # ── JÚNIOR (AO=41, AP=42, AQ=43) ───────────────────────────────────
        sc(ws_f,r,41, "=AO4" if first else 0)
        sc(ws_f,r,42, f"=X{r}-AE{r}-AK{r}" if last else 0)
        sc(ws_f,r,43, f"=-AO{r}+AP{r}")

        # ── Datas auxiliares (AT=46, AU=47, AV=48, AW=49) ──────────────────
        sc(ws_f,r,46, f'=IF(E{r}>0,WORKDAY($D{r}+Dashboard!$I$7-1,1),"")')
        sc(ws_f,r,47, f'=IF(AT{r}<>"",WORKDAY(AT{r}+Dashboard!$L$14-1,1),"")')
        sc(ws_f,r,48, f'=IF(AT{r}<>"",WORKDAY(AT{r}+Dashboard!$L$15-1,1),"")')
        sc(ws_f,r,49, f'=IF(AT{r}<>"",WORKDAY(AT{r}+Dashboard!$L$16-1,1),"")')

        # ── AY (51): Carteira recursiva (para comparação) ──────────────────
        cost_today = (f"SUMIF({rng_au},$D{r},{rng_e})*Dashboard!$J$10"
                      f"+SUMIF({rng_av},$D{r},{rng_j})*Dashboard!$J$11"
                      f"+SUMIF({rng_aw},$D{r},{rng_o})*Dashboard!$J$12")
        if first:
            sc(ws_f,r,51, f"=U{r}")
        else:
            sc(ws_f,r,51, f"=AY{pr}+U{r}-({cost_today})")

        # ── Cores ──────────────────────────────────────────────────────────
        if not last:
            for col in [29, 31, 35, 37, 41, 42]:   # AC, AE, AI, AK, AO, AP
                cell = ws_f.cell(r, col)
                if not isinstance(cell, MergedCell):
                    cell.fill = yellow

        if not wd:
            for col in range(2, 20):   # B até S
                cell = ws_f.cell(r, col)
                if not isinstance(cell, MergedCell):
                    cell.fill = light_red

        # ── Custos ─────────────────────────────────────────────────────────
        sc(ws_c,r,2, n)
        sc(ws_c,r,3, START if first else f"=C{pr}+1")
        sc(ws_c,r,5, "=Dashboard!$J$21/30" if first
                     else f"=MAX(Dashboard!$J$21,(Dashboard!$I$21*Fundo!Z{pr})/360)/30")
        sc(ws_c,r,6, "=Dashboard!$J$20/30" if first
                     else f"=MAX(Dashboard!$J$20,(Dashboard!$I$20*Fundo!Z{pr})/360)/30")
        sc(ws_c,r,7, 0 if first else f"=Fundo!U{pr}*Dashboard!$I$19")
        sc(ws_c,r,8, f"=MAX(0, (Fundo!Y{r} + MAX(0,Fundo!W{r}) - Fundo!Z{pr}*((1+Dashboard!$I$23)^(1/252)-1)) * Dashboard!$I$22)")
        sc(ws_c,r,9, "=Dashboard!I24" if first else None)
        sc(ws_c,r,4, f"=E{r}+F{r}+G{r}+H{r}+I{r}")

    # ── Aba Carteira ───────────────────────────────────────────────────────
    if "Carteira" in wb.sheetnames:
        del wb["Carteira"]
    ws_cart = wb.create_sheet("Carteira")

    orig_days = [n for n in range(1, N + 1) if is_wd(n) and n <= ORIG]

    phases = {}
    for n0 in orig_days:
        dt  = next_wd(n0 + transito)
        dv1 = next_wd(dt + d1)
        dv2 = next_wd(dt + d2)
        dv3 = next_wd(dt + d3)
        phases[n0] = (dt, dv1, dv2, dv3)

    ws_cart.cell(1, 1).value = "Dia"
    ws_cart.cell(1, 2).value = "Data"
    ws_cart.cell(1, 3).value = "Carteira Total"

    fills_cart = [
        PatternFill("solid", fgColor="D6E4F0"),
        PatternFill("solid", fgColor="D5F0D5"),
        PatternFill("solid", fgColor="FFF2CC"),
    ]
    port_qty   = ["E", "J", "O"]
    port_price = ["$J$10", "$J$11", "$J$12"]

    for i, n0 in enumerate(orig_days):
        base_col   = 4 + i*3
        orig_date  = START + datetime.timedelta(days=n0 - 1)
        dt, dv1, dv2, dv3 = phases[n0]
        dv_list = [dv1, dv2, dv3]

        ws_cart.cell(1, base_col).value = f"Orig dia {n0} | {orig_date.strftime('%d/%m/%y')}"

        for p in range(3):
            col  = base_col + p
            dv_p = dv_list[p]

            for n in range(1, N + 1):
                row = n + 3
                if n < n0:
                    val = 0
                elif n < dt:
                    val = f"=Fundo!${port_qty[p]}${n0+4}*Dashboard!{port_price[p]}*Dashboard!$I$6"
                elif n < dv_p:
                    val = f"=Fundo!${port_qty[p]}${n0+4}*Dashboard!{port_price[p]}"
                else:
                    val = 0

                ws_cart.cell(row, col).value = val
                if val != 0:
                    ws_cart.cell(row, col).fill = fills_cart[p]

    # Coluna A (Dia), B (Data), C (Total)
    n_orig = len(orig_days)
    last_col_letter = get_column_letter(4 + n_orig*3 - 1) if n_orig > 0 else "C"
    for n in range(1, N + 1):
        row = n + 3
        ws_cart.cell(row, 1).value = n
        ws_cart.cell(row, 2).value = f"=Fundo!D{n+4}"
        ws_cart.cell(row, 2).number_format = "DD/MM/YY"
        if n_orig > 0:
            ws_cart.cell(row, 3).value = f"=SUM(D{row}:{last_col_letter}{row})"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ════════════════════════════════════════════════════════════════════════════
# DOWNLOAD
# ════════════════════════════════════════════════════════════════════════════

try:
    with open("Modelagem_Frete_v3.xlsx", "rb") as f:
        template_bytes = f.read()
except FileNotFoundError:
    st.error("❌ Arquivo `Modelagem_Frete_v3.xlsx` não encontrado no repositório.")
    st.stop()

with st.spinner("Gerando Excel..."):
    excel_bytes = gerar_excel(template_bytes, n_dias, transito, d1, d2, d3)

st.download_button(
    label=f"⬇️ Baixar Modelagem_Frete_{n_dias}d.xlsx",
    data=excel_bytes,
    file_name=f"Modelagem_Frete_{n_dias}d.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    use_container_width=True,
)
